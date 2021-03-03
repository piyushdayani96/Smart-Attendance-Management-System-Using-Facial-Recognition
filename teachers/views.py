from django.shortcuts import render,redirect,render_to_response
from . import script
from . import detect
from . import attendances
from django.conf import settings
import datetime
import os
from django.core.exceptions import ObjectDoesNotExist
import pyexcel as p
from director.models import Subject
from teachers.models import Studentgroup
from generalzone.models import StudentInfo
import openpyxl
import mimetypes
from django.http import HttpResponse
from wsgiref.util import FileWrapper
from generalzone.models import LoginInfo
from twilio.rest import TwilioRestClient
from django.conf import settings
from django.conf import settings
from django.http import HttpResponse
from twilio.rest import Client
mon=14664
def home(request):
    if request.session['teachid']:
        x=request.session['teachid']
        st=Subject.objects.filter(teachername=x)
        print(st)
        return render(request,'home.html',{'st':st,'x':x})
    else:
        return render(request,'index.html')
def viewstudentgroup(request):
    if request.session['teachid']:
        teachid = request.session['teachid']
        st1 = Studentgroup.objects.filter(teachername=teachid)
        return render(request, 'viewstudentgroup.html', {'st1': st1})
    else:
        return render(request,'index.html')
def createstudentgroup(request,id):
    if request.session['teachid']:
        teachid=request.session['teachid']
        x=id
        print(x)
        return render(request,'createstudentgroup.html',{'x':x})
    else:
        return render(request,'index.html')


def savestudentgroup(request):
    if request.session['teachid']:
        teachid=request.session['teachid']
        studentgroup=request.POST['studentgroup']
        startroll= request.POST['startroll']
        startrolllate= request.POST['startrolllate']
        endroll= request.POST['endroll']
        endrolllate = request.POST['endrolllate']
        st=Studentgroup(studentgroup=studentgroup,teachername=teachid,startroll=startroll,startrolllate=startrolllate,endroll=endroll,endrolllate=endrolllate)
        st.save()
        subjectid = request.POST['subjectid']
        print(subjectid)
        x=Subject.objects.get(id=subjectid)
        x.studentgroup=studentgroup
        x.save(update_fields=["studentgroup"])
        now = datetime.datetime.now()
        today = now.day
        month = now.month
        mn = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
              'November', 'December']
        book = None
        for v in range(0,12):
            if not os.path.isdir("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/"):
               os.makedirs("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/")
            if not os.path.isfile(os.path.join("./media/Attendances/Semester "+str(x.semester)+"/"+str(x.subjectcode)+"/",str(mn[v]) + '.xlsx')):
               book = openpyxl.Workbook()
            else:
               book = openpyxl.load_workbook(os.path.join("./media/Attendances/Semester "+str(x.semester)+"/"+str(x.subjectcode)+"/",str(mn[v]) + '.xlsx'))
            sheet=book.active
            loop=int(endroll)-int(startroll)
            for i in range(0,loop+1):
                try:
                    st=StudentInfo.objects.get(rollno=str(i+int(startroll)))
                    if st is not None:
                        sheet.cell(row=int(i+1),column=2).value=st.name
                    else:
                        sheet.cell(row=int(i + 1), column=2).value ="------"
                except ObjectDoesNotExist:
                    print()
                sheet.cell(row=int(i+1), column=1).value =str(i+int(startroll))
                print('xyz')
                book.save(os.path.join("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/",str(mn[v])+'.xlsx'))
        return redirect('teachers:home')
    else:
        return render(request,'index.html')
def capture(request):
    if request.session['teachid']:
        if request.method == "POST":
            py_obj = script.test_code()
            return render(request,'home.html')
    else:
        return render(request,'index.html')
def face_detect(request):
    if request.session['teachid']:
        if request.method == "POST":
            py = detect.recognize()
            return render(request,'home.html')
    else:
        return render(request,'index.html')
def attend(request):
    if request.session['teachid']:
        if request.method == "POST":
            subjectid = request.POST['attendanceId']
            print(subjectid)
            x=Subject.objects.get(id=subjectid)
            print('pqr')
            p=Studentgroup.objects.get(teachername=x.teachername,studentgroup=x.studentgroup)
            print('mno')
            print(x.semester)
            print(x.subjectcode)
            print(p.startroll)
            print(p.endroll)
            py = attendances.test_attend(x.semester,x.subjectcode,p.startroll,p.endroll)
            st = Subject.objects.filter(teachername=x)
            x.attendanceDateTime = str(str(datetime.datetime.now().strftime("%Y/%m/%d"))+"    "+str(datetime.datetime.now().strftime("%H:%M:%S")))
            x.save(update_fields=['attendanceDateTime'])
            return redirect('teachers:home')
            x = request.session['teachid']

            print(st)
            print("sddfgh")
            print(mon)
            print("asdfg")

            return render(request, 'home.html', {'st': st, 'x': x, 'attendanceDateTime': attendanceDateTime})
    else:
        return render(request,'index.html')
def viewattendance(request):
    if request.session['teachid']:
        now = datetime.datetime.now()
        today = now.day
        month = now.month
        year=now.year
        mn = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
              'November', 'December']
        book = None
        viewId = request.POST['viewAttendance']
        x=Subject.objects.get(id=viewId)
        """if not os.path.isdir("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/"):
           os.makedirs("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/")
        if not os.path.isfile(os.path.join("./media/Attendances/Semester "+str(x.semester)+"/"+str(x.subjectcode)+"/",str(mn[int(month)]) + '.xlsx')):
           book = openpyxl.Workbook()
        else:
           book = openpyxl.load_workbook(os.path.join("./media/Attendances/Semester "+str(x.semester)+"/"+str(x.subjectcode)+"/",str(mn[int(month)]) + '.xlsx'))
        sheet=book.active"""
        wb = openpyxl.load_workbook(os.path.join("./media/Attendances/Semester "+str(x.semester)+"/"+str(x.subjectcode)+"/",str(mn[int(month)-1]) + '.xlsx'))
        worksheet = wb.active
        excel_data = list()
        dates=[]
        for row in worksheet.iter_rows():
            i=1
            for cell in row:
                dates.append(str(i))
                i=i+1
                if(i==33):
                    break
            break
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        st = Subject.objects.filter(id=viewId)
        return render(request, 'home.html', {"excel_data": excel_data,"dates":dates,'st':st})
    else:
        return render(request,'index.html')
def download(request):
    viewId=request.POST['download']
    x = Subject.objects.get(id=viewId)
    print(x.semester)
    print(x.semester)
    files=list()
    mn = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
          'November', 'December']
    t=[0,1,2,3,4,5,6,7,8,9,10,11,12]
    for v in range(0, 12):
        if not os.path.isdir("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/"):
            os.makedirs("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/")
        if os.path.isfile(
            os.path.join("./media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/",
                         str(mn[v]) + '.xlsx')):
            files.append(str("../media/Attendances/Semester " + str(x.semester) + "/" + str(x.subjectcode) + "/"+
                         str(mn[v]) + '.xlsx'))
            print(t[0])
    print(files[0])


    return render(request,'download.html',{'files':files,'x':x,'mn':mn,'t':t})
def awesome_method(request):
    message_to_broadcast = ("Aur bhai kya haal chaal "
                            "yet? Grab it here: https://www.twilio.com/quest")
    client = Client("ACf9cac7ad6f9f4201015286316281e2c7","54e9f910139b89c2b9b9f042e85d4307")
    for recipient in settings.SMS_BROADCAST_TO_NUMBERS:
        if recipient:
            client.messages.create(to='+917985445472',
                                   from_="+12512441061",
                                   body=message_to_broadcast)
    return HttpResponse("messages sent to students with short attendance!", 200)
def indexes(request):
    if request.session['teachid']:
        return render(request,'indexes.html')
    else:
        return render(request,'index.html')





