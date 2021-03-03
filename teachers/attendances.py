import face_recognition as fr
import cv2
import numpy as np
from openpyxl import Workbook
import openpyxl
import datetime
import os
import os
class test_attend:
    def __init__(self,x,y,startroll,endroll):
        print(x)
        images = []
        for filename in os.listdir("./media/images/"):
            images.append(filename)
        # return images
        video_capture = cv2.VideoCapture(0)
        book = None
        """if os.path.isfile('6.xlsx'):
            book = openpyxl.load_workbook('6.xlsx')
        else:
            book = Workbook()"""
        now = datetime.datetime.now()
        today = now.day
        month = now.month
        mn = ['','January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October',
              'November', 'December']
        if not os.path.isdir("./media/Attendances/Semester "+str(x)+"/"+str(y)+"/"):
            os.makedirs("./media/Attendances/Semester "+str(x)+"/"+str(y)+"/")
        if not os.path.isfile(os.path.join("./media/Attendances/Semester "+str(x)+"/"+str(y)+"/",str(mn[int(month)]) + '.xlsx')):
            book=openpyxl.Workbook()
        else:
            book=openpyxl.load_workbook(os.path.join("./media/Attendances/Semester "+str(x)+"/"+str(y)+"/",str(mn[int(month)]) + '.xlsx'))
        sheet = book.active
        images_name = []
        for img in images:
            images_name.append(fr.load_image_file(os.path.join("./media/images/", img)))
        encodings = []
        for img in images_name:
            encodings.append(fr.face_encodings(img)[0])
        # Create array of known face encodings and their names
        known_face_encodings = []
        for encode in encodings:
            known_face_encodings.append(encode)
        known_face_names = []
        for name in images:
            known_face_names.append(os.path.splitext(name)[0])
        # Initialize some variables
        face_locations = []
        face_encodings = []
        face_names = []
        process_this_frame = True
        while True:
            # Grab a single frame of video
            ret, frame = video_capture.read()
            # Resize the frame of video to 1/4 size for fast process
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
            # Convert the image from BGR color(opencv) to RGB color(face_recognition)
            rgb_small_frame = small_frame[:, :, ::-1]
            # Only process every other frame of video to save time
            if process_this_frame:
                # Find all the faces and face encodings in the current frame of video
                face_locations = fr.face_locations(rgb_small_frame)
                face_encodings = fr.face_encodings(rgb_small_frame, face_locations)
                face_names = []
                for face_encoding in face_encodings:
                    # See if the face is a match for known face(s)
                    matches = fr.compare_faces(known_face_encodings, face_encoding)
                    name = "0"
                    face_distances = fr.face_distance(known_face_encodings, face_encoding)
                    best_match_index = np.argmin(face_distances)
                    if matches[best_match_index]:
                        name = known_face_names[best_match_index]
                        if int(name) in range(int(startroll),int(endroll)+1):
                            sheet.cell(row=int(name)-int(startroll)+1, column=int(today)+2).value = "P"
                            if(sheet.cell(row=int(name)-int(startroll)+1,column=35).value is not None):
                                sheet.cell(row=int(name)-int(startroll)+1,column=35).value=int(sheet.cell(row=int(name)-int(startroll)+1,column=35).value)+1
                            else:
                                sheet.cell(row=int(name) - int(startroll) + 1, column=35).value=1;

                        else:
                            pass
                    face_names.append(name)
            process_this_frame = not process_this_frame
            # display the result
            for (top, right, bottom, left), name in zip(face_locations, face_names):
                top *= 4
                right *= 4
                bottom *= 4
                left *= 4
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 225), 2)
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 225), cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (225, 225, 225), 1)
            cv2.imshow('Video', frame)
            book.save(os.path.join('./media/Attendances/Semester '+str(x)+"/"+str(y)+'/',str(mn[int(month)]) + '.xlsx'))
            if cv2.waitKey(1) & 0XFF == ord('q'):
                break
        video_capture.release()
        cv2.destroyAllWindows()




























    """
     def __init__(self):
        # Get a reference to webcam #0 (the default one)
        video_capture = cv2.VideoCapture(0)

        # Create a woorksheet
        book = Workbook()
        sheet = book.active

        # Load images.

        image_1 = face_recognition.load_image_file("./media/images/1.jpg")
        image_1_face_encoding = face_recognition.face_encodings(image_1)[0]

        # Create arrays of known face encodings and their names
        known_face_encodings = [

            image_1_face_encoding,

        ]
        known_face_names = [

          "1"

        ]

        # Initialize some variables
        face_locations = []
        face_encodings = []
        face_names = []
        process_this_frame = True

        # Load present date and time
        now = datetime.datetime.now()
        today = now.day
        month = now.month

        while True:
            # Grab a single frame of video
            ret, frame = video_capture.read()

            # Resize frame of video to 1/4 size for faster face recognition processing
            small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)

            # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
            rgb_small_frame = small_frame[:, :, ::-1]

            # Only process every other frame of video to save time
            if process_this_frame:
                # Find all the faces and face encodings in the current frame of video
                face_locations = face_recognition.face_locations(rgb_small_frame)
                face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []
            for face_encoding in face_encodings:
                # See if the face is a match for the known face(s)
                matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
                name = "Unknown"

                # If a match was found in known_face_encodings, just use the first one.
                if True in matches:
                    first_match_index = matches.index(True)
                    name = known_face_names[first_match_index]
                    # Assign attendance
                    if int(name) in range(1, 61):
                        sheet.cell(row=int(name), column=int(today) + 1).value = "Present"
                    else:
                        pass

                face_names.append(name)

            process_this_frame = not process_this_frame

            # Display the results
                for (top, right, bottom, left), name in zip(face_locations, face_names):
                # Scale back up face locations since the frame we detected in was scaled to 1/4 size
                  top *= 4
                  right *= 4
                  bottom *= 4
                  left *= 4

            # Draw a box around the face
                cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)

            # Draw a label with a name below the face
                cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
                font = cv2.FONT_HERSHEY_DUPLEX
                cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)

            # Display the resulting image
            cv2.imshow('Video', frame)

            # Save Woorksheet as present month
            book.save(str(month) + '.xlsx')

            # Hit 'q' on the keyboard to quit!
            if cv2.waitKey(1) & 0xFF == ord('q'):
                 break

        # Release handle to the webcam
        video_capture.release()
        cv2.destroyAllWindows()


"""